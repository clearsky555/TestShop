from django.http import HttpResponse
from openpyxl import Workbook

from .models import  Product

def get_excel_response():
    products = Product.objects.all()

    wb = Workbook()
    ws = wb.active
    ws.title = 'Products'

    # Заполнение заголовков
    headers = ['Name', 'Price', 'Description']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header

    # Заполнение данных
    for row_num, product in enumerate(products, 2):
        ws.cell(row=row_num, column=1, value=product.name)
        ws.cell(row=row_num, column=2, value=product.price)
        ws.cell(row=row_num, column=3, value=product.description)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=products.xlsx'
    wb.save(response)

    return response
